"""
Tests for ExcelWriterService.

Covers Excel writing, formatting preservation, coloring, and backup creation.
"""
import pytest
from decimal import Decimal
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from src.infrastructure.file_storage.excel_writer import ExcelWriterService
from src.domain.hvac.entities.hvac_description import HVACDescription, HVACDescriptionState
from src.domain.hvac.value_objects.match_score import MatchScore
from src.application.models import ReportFormat


# ============================================================================
# FIXTURES
# ============================================================================


@pytest.fixture
def writer():
    """Fixture for ExcelWriterService instance."""
    return ExcelWriterService()


@pytest.fixture
def temp_excel_file(tmp_path):
    """
    Create a temporary Excel file for testing.

    Structure:
    - Column A: Index (1, 2, 3, ...)
    - Column B: Descriptions
    - Column C: Extra data
    - 5 rows total
    """
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Write header row
    ws["A1"] = "Index"
    ws["B1"] = "Description"
    ws["C1"] = "Extra"

    # Write data rows
    data_rows = [
        (1, "Zawór kulowy DN50 PN16", "Data1"),
        (2, "Zawór zwrotny DN25 PN10", "Data2"),
        (3, "Pompa obiegowa DN32", "Data3"),
        (4, "Filtr siatkowy DN40 PN16", "Data4"),
        (5, "Zawór regulacyjny DN20 PN25", "Data5"),
    ]

    for row_idx, (index, desc, extra) in enumerate(data_rows, start=2):
        ws[f"A{row_idx}"] = index
        ws[f"B{row_idx}"] = desc
        ws[f"C{row_idx}"] = extra

    # Save to temp file
    file_path = tmp_path / "test_file.xlsx"
    wb.save(file_path)

    return file_path


@pytest.fixture
def matched_descriptions():
    """
    Create list of HVACDescription entities with matched prices and scores.

    Returns 3 descriptions:
    - Row 2: High score (95.0) - should be green
    - Row 3: Medium score (80.0) - should be yellow
    - Row 4: Low score (70.0) - should be red
    - Row 5: No match (None price/score)
    """
    # Description 1 - High score, row 2
    desc1 = HVACDescription(
        raw_text="Zawór kulowy DN50 PN16",
        source_row_number=2,
        file_id=uuid4(),
    )
    desc1.match_score = MatchScore.create(
        parameter_score=95.0,
        semantic_score=95.0,
        threshold=75.0,
    )
    desc1.matched_price = Decimal("250.00")
    desc1.matched_description = "Zawór kulowy DN50 PN16 mosiądz"
    desc1.state = HVACDescriptionState.MATCHED

    # Description 2 - Medium score, row 3
    desc2 = HVACDescription(
        raw_text="Zawór zwrotny DN25 PN10",
        source_row_number=3,
        file_id=uuid4(),
    )
    desc2.match_score = MatchScore.create(
        parameter_score=80.0,
        semantic_score=80.0,
        threshold=75.0,
    )
    desc2.matched_price = Decimal("180.00")
    desc2.matched_description = "Zawór zwrotny DN25 PN10"
    desc2.state = HVACDescriptionState.MATCHED

    # Description 3 - Low score, row 4
    desc3 = HVACDescription(
        raw_text="Pompa obiegowa DN32",
        source_row_number=4,
        file_id=uuid4(),
    )
    desc3.match_score = MatchScore.create(
        parameter_score=70.0,
        semantic_score=70.0,
        threshold=75.0,
    )
    desc3.matched_price = Decimal("450.00")
    desc3.matched_description = "Pompa obiegowa DN32 230V"
    desc3.state = HVACDescriptionState.MATCHED

    # Description 4 - No match, row 5
    desc4 = HVACDescription(
        raw_text="Filtr siatkowy DN40 PN16",
        source_row_number=5,
        file_id=uuid4(),
    )
    # No match_score or matched_price

    return [desc1, desc2, desc3, desc4]


# ============================================================================
# TESTS - _column_letter_to_index()
# ============================================================================


def test_column_letter_to_index_single_letter():
    """Test _column_letter_to_index() with single letters (A-Z)."""
    assert ExcelWriterService._column_letter_to_index("A") == 1
    assert ExcelWriterService._column_letter_to_index("B") == 2
    assert ExcelWriterService._column_letter_to_index("C") == 3
    assert ExcelWriterService._column_letter_to_index("Z") == 26


def test_column_letter_to_index_double_letter():
    """Test _column_letter_to_index() with double letters (AA, AB, ...)."""
    assert ExcelWriterService._column_letter_to_index("AA") == 27
    assert ExcelWriterService._column_letter_to_index("AB") == 28
    assert ExcelWriterService._column_letter_to_index("AZ") == 52


def test_column_letter_to_index_case_insensitive():
    """Test that _column_letter_to_index() handles lowercase letters."""
    assert ExcelWriterService._column_letter_to_index("a") == 1
    assert ExcelWriterService._column_letter_to_index("b") == 2
    assert ExcelWriterService._column_letter_to_index("aa") == 27


def test_column_letter_to_index_invalid_input():
    """Test _column_letter_to_index() with invalid input."""
    with pytest.raises(ValueError, match="must contain only letters"):
        ExcelWriterService._column_letter_to_index("A1")

    with pytest.raises(ValueError, match="must contain only letters"):
        ExcelWriterService._column_letter_to_index("")

    with pytest.raises(ValueError, match="must contain only letters"):
        ExcelWriterService._column_letter_to_index("123")


# ============================================================================
# TESTS - _create_backup()
# ============================================================================


def test_create_backup_success(writer, temp_excel_file):
    """Test _create_backup() creates backup file successfully."""
    backup_path = writer._create_backup(temp_excel_file)

    # Check backup path format
    expected_backup = temp_excel_file.parent / f"{temp_excel_file.stem}_backup{temp_excel_file.suffix}"
    assert backup_path == expected_backup

    # Check backup file exists
    assert backup_path.exists()

    # Check backup content matches original
    original_wb = load_workbook(temp_excel_file)
    backup_wb = load_workbook(backup_path)

    assert original_wb.active["B2"].value == backup_wb.active["B2"].value


def test_create_backup_file_not_found(writer, tmp_path):
    """Test _create_backup() with non-existent file."""
    nonexistent_file = tmp_path / "nonexistent.xlsx"

    with pytest.raises(FileNotFoundError, match="File not found"):
        writer._create_backup(nonexistent_file)


# ============================================================================
# TESTS - _load_workbook() & _get_worksheet()
# ============================================================================


def test_load_workbook_success(writer, temp_excel_file):
    """Test _load_workbook() loads Excel file successfully."""
    wb = writer._load_workbook(temp_excel_file)

    assert isinstance(wb, Workbook)
    assert "Sheet1" in wb.sheetnames


def test_load_workbook_file_not_found(writer, tmp_path):
    """Test _load_workbook() with non-existent file."""
    nonexistent_file = tmp_path / "nonexistent.xlsx"

    with pytest.raises(FileNotFoundError, match="File not found"):
        writer._load_workbook(nonexistent_file)


def test_get_worksheet_default_sheet(writer, temp_excel_file):
    """Test _get_worksheet() returns active sheet when sheet_name is None."""
    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, None)

    assert ws.title == "Sheet1"


def test_get_worksheet_named_sheet(writer, temp_excel_file):
    """Test _get_worksheet() returns named sheet."""
    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, "Sheet1")

    assert ws.title == "Sheet1"


def test_get_worksheet_sheet_not_found(writer, temp_excel_file):
    """Test _get_worksheet() with non-existent sheet name."""
    wb = writer._load_workbook(temp_excel_file)

    with pytest.raises(ValueError, match="Sheet 'NonExistent' not found"):
        writer._get_worksheet(wb, "NonExistent")


# ============================================================================
# TESTS - _write_prices_to_column()
# ============================================================================


def test_write_prices_to_column_success(writer, temp_excel_file, matched_descriptions):
    """Test _write_prices_to_column() writes prices correctly."""
    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, None)

    # Write prices to column D
    writer._write_prices_to_column(ws, matched_descriptions, "D")

    # Check prices written (rows 2, 3, 4 have prices)
    assert ws["D2"].value == 250.00  # desc1
    assert ws["D3"].value == 180.00  # desc2
    assert ws["D4"].value == 450.00  # desc3
    assert ws["D5"].value is None    # desc4 - no match


def test_write_prices_skips_empty_prices(writer, temp_excel_file):
    """Test _write_prices_to_column() skips descriptions without prices."""
    # Create description without matched_price
    desc = HVACDescription(
        raw_text="Test",
        source_row_number=2,
        file_id=uuid4(),
    )
    # No matched_price set

    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, None)

    writer._write_prices_to_column(ws, [desc], "D")

    # Cell should remain empty
    assert ws["D2"].value is None


# ============================================================================
# TESTS - _apply_cell_coloring()
# ============================================================================


def test_apply_cell_coloring_green_yellow_red(writer, temp_excel_file, matched_descriptions):
    """Test _apply_cell_coloring() applies correct colors based on scores."""
    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, None)

    # First write prices (coloring needs prices to be present)
    writer._write_prices_to_column(ws, matched_descriptions, "D")

    # Apply coloring
    writer._apply_cell_coloring(ws, matched_descriptions, "D")

    # Check colors (openpyxl uses 00RRGGBB format, not FFRRGGBB)
    # Row 2: score 95 -> green (00FF00)
    green_rgb = ws["D2"].fill.start_color.rgb
    assert green_rgb[-6:] == "00FF00"  # Check last 6 chars (RRGGBB)

    # Row 3: score 80 -> yellow (FFFF00)
    yellow_rgb = ws["D3"].fill.start_color.rgb
    assert yellow_rgb[-6:] == "FFFF00"

    # Row 4: score 70 -> red (FF0000)
    red_rgb = ws["D4"].fill.start_color.rgb
    assert red_rgb[-6:] == "FF0000"

    # Row 5: no match -> no color (default fill)
    default_rgb = ws["D5"].fill.start_color.rgb
    assert default_rgb == "00000000"  # Default


# ============================================================================
# TESTS - _get_color_for_score()
# ============================================================================


def test_get_color_for_score_green():
    """Test _get_color_for_score() returns green for score > 90."""
    writer = ExcelWriterService()
    assert writer._get_color_for_score(95.0) == "00FF00"
    assert writer._get_color_for_score(91.0) == "00FF00"


def test_get_color_for_score_yellow():
    """Test _get_color_for_score() returns yellow for score 75-90."""
    writer = ExcelWriterService()
    assert writer._get_color_for_score(90.0) == "FFFF00"
    assert writer._get_color_for_score(80.0) == "FFFF00"
    assert writer._get_color_for_score(75.0) == "FFFF00"


def test_get_color_for_score_red():
    """Test _get_color_for_score() returns red for score < 75."""
    writer = ExcelWriterService()
    assert writer._get_color_for_score(74.9) == "FF0000"
    assert writer._get_color_for_score(50.0) == "FF0000"
    assert writer._get_color_for_score(0.0) == "FF0000"


# ============================================================================
# TESTS - _autosize_columns()
# ============================================================================


def test_autosize_columns_adjusts_width(writer, temp_excel_file):
    """Test _autosize_columns() adjusts column widths."""
    wb = writer._load_workbook(temp_excel_file)
    ws = writer._get_worksheet(wb, None)

    # Auto-size column B (has descriptions)
    writer._autosize_columns(ws, ["B"])

    # Width should be adjusted (greater than default)
    # Longest description in B: "Zawór regulacyjny DN20 PN25" (28 chars)
    # Expected width: 28 + 2 = 30
    assert ws.column_dimensions["B"].width >= 28


# ============================================================================
# TESTS - write_results_to_file() - INTEGRATION
# ============================================================================


def test_write_results_to_file_success(writer, temp_excel_file, matched_descriptions, tmp_path):
    """Integration test: write_results_to_file() creates result file successfully."""
    output_path = tmp_path / "result.xlsx"

    # Write results
    result_path = writer.write_results_to_file(
        original_file_path=temp_excel_file,
        descriptions=matched_descriptions,
        price_column="D",
        report_column=None,
        output_path=output_path,
    )

    # Check result file created
    assert result_path == output_path
    assert result_path.exists()

    # Check backup created
    backup_path = temp_excel_file.parent / f"{temp_excel_file.stem}_backup{temp_excel_file.suffix}"
    assert backup_path.exists()

    # Check prices written
    result_wb = load_workbook(result_path)
    result_ws = result_wb.active

    assert result_ws["D2"].value == 250.00
    assert result_ws["D3"].value == 180.00
    assert result_ws["D4"].value == 450.00
    assert result_ws["D5"].value is None

    # Check coloring applied (check last 6 chars of RGB)
    assert result_ws["D2"].fill.start_color.rgb[-6:] == "00FF00"  # Green
    assert result_ws["D3"].fill.start_color.rgb[-6:] == "FFFF00"  # Yellow
    assert result_ws["D4"].fill.start_color.rgb[-6:] == "FF0000"  # Red


def test_write_results_to_file_with_reports(writer, temp_excel_file, matched_descriptions, tmp_path):
    """Integration test: write_results_to_file() writes prices and reports."""
    output_path = tmp_path / "result_with_reports.xlsx"

    # Write results with reports
    result_path = writer.write_results_to_file(
        original_file_path=temp_excel_file,
        descriptions=matched_descriptions,
        price_column="D",
        report_column="E",
        output_path=output_path,
        report_format=ReportFormat.SIMPLE,
    )

    # Check result file created
    assert result_path.exists()

    # Check reports written
    result_wb = load_workbook(result_path)
    result_ws = result_wb.active

    # Rows with matches should have reports (2, 3, 4)
    # Row 5 has no match, so no report
    assert result_ws["E2"].value is not None  # Has report
    assert result_ws["E3"].value is not None
    assert result_ws["E4"].value is not None
    assert result_ws["E5"].value is None      # No match, no report


def test_write_results_to_file_default_output_path(writer, temp_excel_file, matched_descriptions):
    """Test write_results_to_file() uses default output path when not specified."""
    # Don't specify output_path
    result_path = writer.write_results_to_file(
        original_file_path=temp_excel_file,
        descriptions=matched_descriptions,
        price_column="D",
    )

    # Should create result.xlsx in same directory as original
    expected_path = temp_excel_file.parent / "result.xlsx"
    assert result_path == expected_path
    assert result_path.exists()


def test_write_results_to_file_preserves_formatting(writer, temp_excel_file, matched_descriptions, tmp_path):
    """Test write_results_to_file() preserves original Excel formatting."""
    output_path = tmp_path / "result_formatted.xlsx"

    # Write results
    writer.write_results_to_file(
        original_file_path=temp_excel_file,
        descriptions=matched_descriptions,
        price_column="D",
        output_path=output_path,
    )

    # Check original data preserved
    result_wb = load_workbook(output_path)
    result_ws = result_wb.active

    # Original columns should be intact
    assert result_ws["A1"].value == "Index"
    assert result_ws["B1"].value == "Description"
    assert result_ws["C1"].value == "Extra"
    assert result_ws["B2"].value == "Zawór kulowy DN50 PN16"
    assert result_ws["C2"].value == "Data1"
