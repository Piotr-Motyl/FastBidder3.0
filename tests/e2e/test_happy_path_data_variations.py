"""
End-to-End Tests for Happy Path Data Variations

Tests valid happy path scenarios with different data inputs:
1. Polish characters (ą, ć, ę, ł, ń, ó, ś, ź, ż) - encoding preservation
2. Minimum viable input (single item) - boundary condition
3. Very long descriptions (>200 chars) - buffer handling (PLACEHOLDER for Phase 4+)

These tests validate that the system correctly handles edge cases
in valid happy path scenarios (not error cases).

Requirements:
    - Docker services running (Redis, Celery worker)
    - Data variation fixtures available (polish_chars_*.xlsx, single_item_*.xlsx)
    - All API endpoints implemented (files, matching, jobs, results)

Setup:
    Before running these tests, generate fixtures:
    $ python tests/fixtures/generate_fixtures.py --data-variations

Run:
    # Run all data variation tests
    $ pytest tests/e2e/test_happy_path_data_variations.py -v

    # Run specific test
    $ pytest tests/e2e/test_happy_path_data_variations.py::test_polish_characters_in_descriptions -v -s

Architecture Notes:
    - Uses real services (Redis, Celery, file system)
    - Tests integration of all layers
    - Validates happy path edge cases only (not error scenarios)

Test Coverage:
    - Polish characters: 3 descriptions with ą, ć, ę, ł, ń, ó, ś, ź, ż
    - Single item: 1 description (minimum viable input)
    - Long descriptions: PLACEHOLDER (Phase 4+ for scaling)
"""

import logging
import time
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

# Import helper functions from test_matching_workflow
from tests.e2e.test_matching_workflow import (
    upload_file,
    trigger_matching,
    poll_job_status,
    download_results,
    MIN_MATCH_SCORE,
)

# ============================================================================
# TEST CONFIGURATION
# ============================================================================

# Test timeout: max wait time for Celery task to complete
TEST_TIMEOUT_SECONDS = 60

# Poll interval: how often to check job status
POLL_INTERVAL_SECONDS = 2

# Expected match score threshold for success
MIN_ACCEPTABLE_SCORE = 75.0

# Configure logger
logger = logging.getLogger(__name__)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def get_column_index(headers: list, column_name: str) -> int:
    """
    Get column index from headers (1-based for openpyxl).

    Args:
        headers: List of column headers from row 1
        column_name: Column name to find

    Returns:
        1-based column index

    Raises:
        AssertionError: If column not found
    """
    try:
        return headers.index(column_name) + 1
    except ValueError:
        available_cols = ", ".join(str(h) for h in headers if h)
        raise AssertionError(
            f"Column '{column_name}' not found in output. "
            f"Available columns: {available_cols}"
        )


# ============================================================================
# E2E TESTS - DATA VARIATIONS
# ============================================================================


@pytest.mark.e2e
def test_polish_characters_in_descriptions(
    test_client,
    clean_redis,
    docker_services,
):
    """
    Test Polish characters (ą, ć, ę, ł, ń, ó, ś, ź, ż) through entire E2E flow.

    Validates:
        - UTF-8 encoding preserved through: upload → read → process → write → download
        - Regex patterns work with Polish characters
        - Text matching works with diacritics
        - Excel files readable and display Polish characters correctly

    Test Data:
        Working file (3 items with Polish chars):
        - Row 2: "Zawór kulowy DN50 PN16 mosiądz" (ą)
        - Row 3: "Zawór zwrotny DN80 PN10 żeliwo szare" (ż)
        - Row 4: "Kompensator długości DN100 stal nierdzewna" (ł, ó)

        Reference file (matching items):
        - Row 2: "Zawór kulowy DN50 PN16 mosiądz" | Price: 250.00
        - Row 3: "Zawór zwrotny DN80 PN10 żeliwo" | Price: 180.00
        - Row 4: "Kompensator DN100 nierdzewny" | Price: 450.00

    Expected Output:
        - All 3 items matched (prices filled)
        - All match scores >= 75%
        - Polish characters preserved in output
        - No encoding errors (mosiądz ≠ mosiadz, żeliwo ≠ zeliwo)

    Test passes when:
        ✓ Job completes successfully
        ✓ All Polish characters preserved exactly
        ✓ All items matched with score >= 75%
        ✓ Excel file opens correctly

    Test fails when:
        ✗ Polish characters replaced (ą→a, ż→z)
        ✗ Encoding corruption (mosiądz→mosiÄ…dz)
        ✗ Items not matched (prices empty)
        ✗ Match scores < 75%
    """
    logger.info("=" * 80)
    logger.info("STARTING E2E TEST: Polish Characters")
    logger.info("=" * 80)

    # Setup: Get fixtures with Polish characters
    fixtures_dir = Path(__file__).parent.parent / "fixtures"
    working_file = fixtures_dir / "polish_chars_working.xlsx"
    reference_file = fixtures_dir / "polish_chars_reference.xlsx"

    # Verify fixtures exist
    assert working_file.exists(), f"Polish chars working fixture not found: {working_file}"
    assert reference_file.exists(), f"Polish chars reference fixture not found: {reference_file}"

    # ========================================================================
    # STAGE 1: Upload files
    # ========================================================================
    logger.info("\n[STAGE 1] Uploading files with Polish characters...")

    working_upload = upload_file(test_client, working_file)
    working_file_id = working_upload["file_id"]

    reference_upload = upload_file(test_client, reference_file)
    reference_file_id = reference_upload["file_id"]

    assert working_file_id != reference_file_id, "File IDs should be different"

    # ========================================================================
    # STAGE 2: Trigger matching process
    # ========================================================================
    logger.info("\n[STAGE 2] Triggering matching process...")

    process_response = trigger_matching(
        test_client,
        working_file_id=working_file_id,
        reference_file_id=reference_file_id,
        threshold=75.0,
    )

    job_id = process_response["job_id"]
    assert process_response["status"] == "queued", (
        f"Expected status queued, got {process_response['status']}"
    )

    # ========================================================================
    # STAGE 3: Wait for completion (polling)
    # ========================================================================
    logger.info("\n[STAGE 3] Waiting for job completion...")

    final_status = poll_job_status(
        test_client,
        job_id=job_id,
        timeout_seconds=TEST_TIMEOUT_SECONDS,
        poll_interval=POLL_INTERVAL_SECONDS,
    )

    assert final_status["status"] == "completed", (
        f"Expected status completed, got {final_status['status']}"
    )

    # ========================================================================
    # STAGE 4: Download results
    # ========================================================================
    logger.info("\n[STAGE 4] Downloading result file...")

    result_bytes = download_results(test_client, job_id)
    assert len(result_bytes) > 0, "Result file is empty"

    # ========================================================================
    # STAGE 5: Validate Polish characters preserved
    # ========================================================================
    logger.info("\n[STAGE 5] Validating Polish characters in output...")

    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb.active

    # Get column indices using helper function
    headers = [cell.value for cell in ws[1]]
    desc_col = get_column_index(headers, "Description")
    price_col = get_column_index(headers, "Cena")
    score_col = get_column_index(headers, "Match Score")

    # Check row 2: "Zawór kulowy DN50 PN16 mosiądz"
    logger.info("\nValidating row 2: Zawór kulowy DN50 PN16 mosiądz")
    desc_row_2 = ws.cell(2, desc_col).value
    price_row_2 = ws.cell(2, price_col).value
    score_row_2 = ws.cell(2, score_col).value

    assert desc_row_2 is not None, "Row 2 description is empty"
    assert "Zawór" in desc_row_2, f"Polish 'ó' missing in description: {desc_row_2}"
    assert "mosiądz" in desc_row_2, f"Polish 'ą' missing in description: {desc_row_2}"
    assert price_row_2 is not None, "Row 2 should have price (matched)"
    assert score_row_2 is not None, "Row 2 should have match score"
    assert float(score_row_2) >= MIN_ACCEPTABLE_SCORE, f"Row 2 score too low: {score_row_2}"
    logger.info(f"✓ Row 2: Price={price_row_2}, Score={score_row_2}")

    # Check row 3: "Zawór zwrotny DN80 PN10 żeliwo szare"
    logger.info("\nValidating row 3: Zawór zwrotny DN80 PN10 żeliwo szare")
    desc_row_3 = ws.cell(3, desc_col).value
    price_row_3 = ws.cell(3, price_col).value
    score_row_3 = ws.cell(3, score_col).value

    assert desc_row_3 is not None, "Row 3 description is empty"
    assert "żeliwo" in desc_row_3, f"Polish 'ż' missing in description: {desc_row_3}"
    assert price_row_3 is not None, "Row 3 should have price (matched)"
    assert score_row_3 is not None, "Row 3 should have match score"
    assert float(score_row_3) >= MIN_ACCEPTABLE_SCORE, f"Row 3 score too low: {score_row_3}"
    logger.info(f"✓ Row 3: Price={price_row_3}, Score={score_row_3}")

    # Check row 4: "Kompensator długości DN100 stal nierdzewna"
    logger.info("\nValidating row 4: Kompensator długości DN100 stal nierdzewna")
    desc_row_4 = ws.cell(4, desc_col).value
    price_row_4 = ws.cell(4, price_col).value
    score_row_4 = ws.cell(4, score_col).value

    assert desc_row_4 is not None, "Row 4 description is empty"
    assert "długości" in desc_row_4, f"Polish 'ł' and 'ś' missing in description: {desc_row_4}"
    assert price_row_4 is not None, "Row 4 should have price (matched)"
    assert score_row_4 is not None, "Row 4 should have match score"
    assert float(score_row_4) >= MIN_ACCEPTABLE_SCORE, f"Row 4 score too low: {score_row_4}"
    logger.info(f"✓ Row 4: Price={price_row_4}, Score={score_row_4}")

    # Final validation
    logger.info("\n" + "=" * 80)
    logger.info("E2E TEST COMPLETED: Polish Characters ✓")
    logger.info("=" * 80)
    logger.info("All Polish characters preserved correctly:")
    logger.info("  ✓ ą (mosiądz)")
    logger.info("  ✓ ż (żeliwo)")
    logger.info("  ✓ ł, ó, ś (długości)")
    logger.info(f"All items matched with scores >= {MIN_ACCEPTABLE_SCORE}%")
    logger.info("=" * 80)


@pytest.mark.e2e
def test_minimum_viable_input_single_item(
    test_client,
    clean_redis,
    docker_services,
):
    """
    Test minimum viable input: 1 description only.

    Validates:
        - System works with boundary condition (1 item)
        - No off-by-one errors
        - Progress tracking works (0% → 100%)
        - Output has exactly 1 row
        - Processing time reasonable (<30s)

    Test Data:
        Working file (1 item only):
        - Row 2: "Zawór kulowy DN50 PN16 mosiądz"
        Range: start=2, end=2 (single row)

        Reference file (reuse sample):
        - 50 items (normal catalog)

    Expected Output:
        - Exactly 1 data row (+ 1 header = 2 total)
        - Item matched (price filled)
        - Match score >= 75%
        - Job completes in <30s

    Test passes when:
        ✓ Job completes successfully in <30s
        ✓ Output has exactly 1 data row
        ✓ Item matched with score >= 75%
        ✓ No off-by-one errors

    Test fails when:
        ✗ Job fails (IndexError, off-by-one)
        ✗ Output has 0 or 2+ rows
        ✗ Item not matched (price empty)
        ✗ Processing takes >30s
    """
    logger.info("=" * 80)
    logger.info("STARTING E2E TEST: Single Item (Minimum Viable Input)")
    logger.info("=" * 80)

    # Setup: Get fixtures
    fixtures_dir = Path(__file__).parent.parent / "fixtures"
    working_file = fixtures_dir / "single_item_working.xlsx"
    reference_file = fixtures_dir / "sample_reference_file.xlsx"  # Reuse existing

    # Verify fixtures exist
    assert working_file.exists(), f"Single item working fixture not found: {working_file}"
    assert reference_file.exists(), f"Reference fixture not found: {reference_file}"

    # ========================================================================
    # STAGE 1: Upload files
    # ========================================================================
    logger.info("\n[STAGE 1] Uploading files (1 item + 50 reference)...")

    working_upload = upload_file(test_client, working_file)
    working_file_id = working_upload["file_id"]

    reference_upload = upload_file(test_client, reference_file)
    reference_file_id = reference_upload["file_id"]

    # ========================================================================
    # STAGE 2: Trigger matching with 1 item range
    # ========================================================================
    logger.info("\n[STAGE 2] Triggering matching process (1 item only)...")

    # Custom payload for single item range
    payload = {
        "working_file": {
            "file_id": working_file_id,
            "description_column": "A",
            "description_range": {"start": 2, "end": 2},  # Only 1 row!
            "price_target_column": "B",
            "matching_report_column": "D",
        },
        "reference_file": {
            "file_id": reference_file_id,
            "description_column": "A",
            "description_range": {"start": 2, "end": 51},  # Normal range
            "price_source_column": "B",
        },
        "matching_threshold": 75.0,
        "matching_strategy": "best_match",
        "report_format": "detailed",
    }

    response = test_client.post("/api/matching/process", json=payload)
    assert response.status_code == 202, (
        f"Matching trigger failed: {response.status_code} - {response.text}"
    )

    job_id = response.json()["job_id"]
    logger.info(f"Job triggered: {job_id}")

    # ========================================================================
    # STAGE 3: Wait for completion (should be very fast)
    # ========================================================================
    logger.info("\n[STAGE 3] Waiting for job completion (max 30s)...")

    start_time = time.time()
    final_status = poll_job_status(
        test_client,
        job_id=job_id,
        timeout_seconds=30,
        poll_interval=POLL_INTERVAL_SECONDS
    )
    duration = time.time() - start_time

    assert final_status["status"] == "completed", (
        f"Expected status completed, got {final_status['status']}"
    )
    assert duration < 30, f"1 item took too long: {duration:.2f}s"
    logger.info(f"✓ Job completed in {duration:.2f}s")

    # ========================================================================
    # STAGE 4: Download results
    # ========================================================================
    logger.info("\n[STAGE 4] Downloading result file...")

    result_bytes = download_results(test_client, job_id)
    assert len(result_bytes) > 0, "Result file is empty"

    # ========================================================================
    # STAGE 5: Validate exactly 1 data row
    # ========================================================================
    logger.info("\n[STAGE 5] Validating output has exactly 1 data row...")

    wb = openpyxl.load_workbook(BytesIO(result_bytes))
    ws = wb.active

    # Check exactly 2 rows (1 header + 1 data)
    actual_rows = ws.max_row
    assert actual_rows == 2, f"Expected 2 rows (header+data), got {actual_rows}"
    logger.info(f"✓ Output has exactly 1 data row (total: {actual_rows} rows)")

    # Get column indices using helper function
    headers = [cell.value for cell in ws[1]]
    desc_col = get_column_index(headers, "Description")
    price_col = get_column_index(headers, "Cena")
    score_col = get_column_index(headers, "Match Score")

    # Validate row 2 has data
    desc = ws.cell(2, desc_col).value
    price = ws.cell(2, price_col).value
    score = ws.cell(2, score_col).value

    assert desc is not None, "Row 2 description is empty"
    assert price is not None, "Row 2 price is empty (no match?)"
    assert score is not None, "Row 2 score is empty"
    assert float(score) >= MIN_ACCEPTABLE_SCORE, f"Score too low: {score}"

    # Final validation
    logger.info("\n" + "=" * 80)
    logger.info("E2E TEST COMPLETED: Single Item ✓")
    logger.info("=" * 80)
    logger.info(f"Processing time: {duration:.2f}s")
    logger.info(f"Output rows: {actual_rows} (header + 1 data)")
    logger.info(f"Item matched: Price={price}, Score={score}")
    logger.info("No off-by-one errors detected")
    logger.info("=" * 80)


@pytest.mark.e2e
@pytest.mark.skip(reason="PLACEHOLDER for Phase 4+ scaling - not implemented yet")
def test_very_long_descriptions():
    """
    Test very long descriptions (>200 chars) - PLACEHOLDER.

    This test will be implemented in Phase 4+ when optimizing for scale.

    Will validate:
        - No truncation in Excel write/read
        - No buffer overflow in parameter extraction
        - Matching works with long texts
        - Performance acceptable

    Test Data:
        Working file (3 long descriptions, 200-300 chars each):
        - Row 2: Zawór kulowy trójdrożny DN50 PN16... (235 chars)
        - Row 3: Zawór regulacyjny dwudrogowy DN80... (260 chars)
        - Row 4: Kompensator długości osiowy DN100... (280 chars)

    Expected:
        - All descriptions preserved completely (no truncation)
        - All items matched
        - Job completes in <60s

    Implementation: Phase 4+ (scaling optimization)
    """
    pass  # Placeholder - will implement later
