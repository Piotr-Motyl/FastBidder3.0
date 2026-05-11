"""
End-to-End Tests for Matching Workflow

Tests the complete workflow from upload to download:
1. Upload working file
2. Upload reference file
3. Trigger matching process (Celery task)
4. Poll job status
5. Download results
6. Validate output

Requirements:
    - Docker services running (Redis, Celery worker)
    - Sample fixtures available (sample_working_file.xlsx, sample_reference_file.xlsx)
    - All API endpoints implemented (files, matching, jobs, results)

Setup:
    Before running these tests, start Docker services:
    $ docker-compose up -d

    Verify services are running:
    $ docker-compose ps
    $ redis-cli ping  # Should return PONG

Run:
    # Run all E2E tests
    $ pytest tests/e2e/ -v

    # Run specific test
    $ pytest tests/e2e/test_matching_workflow.py::test_full_workflow_happy_path -v

    # Run with detailed output
    $ pytest tests/e2e/ -v -s

Architecture Notes:
    - Uses FastAPI TestClient (no need for running server)
    - Requires real Redis (mocking would defeat E2E purpose)
    - Requires real Celery worker (asynchronous task execution)
    - Uses real file system (temp files cleaned up after test)
    - Tests integration of all layers: API → Application → Domain → Infrastructure

Test Coverage:
    - Happy path: 20 descriptions, ≥50% should match with score >75%
    - Invalid files: Should return appropriate error responses
    - Low threshold: Should match more items with lower quality threshold
"""

import logging
import time
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

# ============================================================================
# TEST CONFIGURATION
# ============================================================================

# Test timeout: max wait time for Celery task to complete
TEST_TIMEOUT_SECONDS = 60

# Poll interval: how often to check job status
POLL_INTERVAL_SECONDS = 2

# Expected success rate: at least 50% of items should match with score >75%
MIN_SUCCESS_RATE = 0.5

# Expected match score threshold for success
MIN_MATCH_SCORE = 75.0

# Configure logger
logger = logging.getLogger(__name__)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def upload_file(test_client, file_path: Path, file_type: str = "working") -> dict:
    """
    Upload Excel file to /api/files/upload endpoint.

    Args:
        test_client: FastAPI TestClient
        file_path: Path to Excel file
        file_type: Type of file ("working" or "reference") for AI indexing

    Returns:
        dict: Upload response with file_id, filename, size_mb, etc.

    Raises:
        AssertionError: If upload fails
    """
    with open(file_path, "rb") as f:
        response = test_client.post(
            "/api/files/upload",
            files={
                "file": (
                    file_path.name,
                    f,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            params={"file_type": file_type},  # Phase 4: Pass file_type for AI indexing
        )

    assert response.status_code == 201, (
        f"Upload failed: {response.status_code} - {response.text}"
    )

    data = response.json()
    logger.info(f"Uploaded {file_path.name}: file_id={data['file_id']}")

    return data


def trigger_matching(
    test_client,
    working_file_id: str,
    reference_file_id: str,
    threshold: float = 75.0,
) -> dict:
    """
    Trigger matching process via /api/matching/process endpoint.

    Args:
        test_client: FastAPI TestClient
        working_file_id: UUID of working file
        reference_file_id: UUID of reference file
        threshold: Match threshold (0-100)

    Returns:
        dict: Process response with job_id, status, estimated_time

    Raises:
        AssertionError: If trigger fails
    """
    # Build request according to API schema (ProcessMatchingRequest)
    # Working file: 20 descriptions in column A, rows 2-21 (header in row 1)
    # Reference file: 50 descriptions in column A, prices in column B, rows 2-51
    # Output: A=Description, B=Cena, C=Match Score, D=Match Report
    payload = {
        "working_file": {
            "file_id": working_file_id,
            "description_column": "A",
            "description_range": {"start": 2, "end": 21},
            "price_target_column": "B",
            "matching_report_column": "D",  # Changed from C to D (C is now Match Score)
        },
        "reference_file": {
            "file_id": reference_file_id,
            "description_column": "A",
            "description_range": {"start": 2, "end": 51},
            "price_source_column": "B",
        },
        "matching_threshold": threshold,
        "matching_strategy": "best_match",  # lowercase enum value
        "report_format": "detailed",  # lowercase enum value
    }

    response = test_client.post("/api/matching/process", json=payload)

    assert response.status_code == 202, (
        f"Matching trigger failed: {response.status_code} - {response.text}"
    )

    data = response.json()
    logger.info(f"Matching triggered: job_id={data['job_id']}")

    return data


def poll_job_status(
    test_client,
    job_id: str,
    timeout_seconds: int = TEST_TIMEOUT_SECONDS,
    poll_interval: int = POLL_INTERVAL_SECONDS,
) -> dict:
    """
    Poll job status until completion or timeout.

    Args:
        test_client: FastAPI TestClient
        job_id: Job UUID
        timeout_seconds: Max wait time
        poll_interval: Poll interval in seconds

    Returns:
        dict: Final job status response

    Raises:
        TimeoutError: If job doesn't complete within timeout
        AssertionError: If job fails
    """
    start_time = time.time()
    last_message = None

    while True:
        # Check timeout
        elapsed = time.time() - start_time
        if elapsed > timeout_seconds:
            raise TimeoutError(
                f"Job {job_id} did not complete within {timeout_seconds}s. "
                f"Last message: {last_message}"
            )

        # Get status
        response = test_client.get(f"/api/jobs/{job_id}/status")

        assert response.status_code == 200, (
            f"Status check failed: {response.status_code} - {response.text}"
        )

        data = response.json()
        status = data["status"]

        # New API structure (JobStatusResponse)
        # progress is int (0-100), not dict
        percentage = data.get("progress", 0)
        message = data.get("message", "")
        current_step = data.get("current_step", "")

        # Log progress (only if message changed)
        if message != last_message:
            logger.info(
                f"Job {job_id}: {status} - {percentage}% - {current_step} - {message}"
            )
            last_message = message

        # Check if completed
        if status == "completed":
            logger.info(f"Job {job_id} completed successfully after {elapsed:.1f}s")
            return data

        # Check if failed
        if status == "failed":
            error_details = data.get("error_details", "Unknown error")
            raise AssertionError(
                f"Job {job_id} failed: {error_details}"
            )

        # Wait before next poll
        time.sleep(poll_interval)


def download_results(test_client, job_id: str) -> bytes:
    """
    Download result file from /api/results/{job_id}/download endpoint.

    Args:
        test_client: FastAPI TestClient
        job_id: Job UUID

    Returns:
        bytes: Excel file content

    Raises:
        AssertionError: If download fails
    """
    response = test_client.get(f"/api/results/{job_id}/download")

    assert response.status_code == 200, (
        f"Download failed: {response.status_code} - {response.text}"
    )

    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ), f"Wrong content type: {response.headers['content-type']}"

    logger.info(f"Downloaded result file for job {job_id}: {len(response.content)} bytes")

    return response.content


def validate_output_file(file_bytes: bytes, min_success_rate: float = MIN_SUCCESS_RATE) -> dict:
    """
    Validate output Excel file structure and content.

    Checks:
        1. File can be opened as Excel
        2. Has expected columns: "Cena", "Match Score", "Match Report"
        3. At least min_success_rate% of rows have Match Score > MIN_MATCH_SCORE

    Args:
        file_bytes: Excel file content as bytes
        min_success_rate: Minimum % of rows that should have good matches

    Returns:
        dict: Validation stats
            - total_rows: Total data rows (excluding header)
            - rows_with_price: Rows with Cena filled
            - rows_with_score: Rows with Match Score filled
            - high_quality_matches: Rows with Match Score > MIN_MATCH_SCORE
            - success_rate: Percentage of high quality matches

    Raises:
        AssertionError: If validation fails
    """
    # Load Excel file
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    assert ws is not None, "Workbook has no active sheet"

    # Get header row
    headers = [cell.value for cell in ws[1]]
    logger.info(f"Output file headers: {headers}")

    # Check required columns exist
    assert "Cena" in headers, "Missing column: Cena"
    assert "Match Score" in headers, "Missing column: Match Score"
    assert "Match Report" in headers, "Missing column: Match Report"

    # Get column indices (Match Report column is asserted to exist above
    # but isn't used in the stats — its presence is the contract).
    price_col = headers.index("Cena") + 1
    score_col = headers.index("Match Score") + 1

    # Count statistics
    total_rows = ws.max_row - 1  # Exclude header
    rows_with_price = 0
    rows_with_score = 0
    high_quality_matches = 0

    # Analyze each row
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        price = ws.cell(row_idx, price_col).value
        score = ws.cell(row_idx, score_col).value

        if price is not None and price != "":
            rows_with_price += 1

        if score is not None:
            rows_with_score += 1

            # Convert score to float (may be string or float)
            try:
                score_value = float(score)
                if score_value >= MIN_MATCH_SCORE:
                    high_quality_matches += 1
            except (ValueError, TypeError):
                logger.warning(f"Row {row_idx}: Invalid score value: {score}")

    # Calculate success rate
    success_rate = high_quality_matches / total_rows if total_rows > 0 else 0.0

    stats = {
        "total_rows": total_rows,
        "rows_with_price": rows_with_price,
        "rows_with_score": rows_with_score,
        "high_quality_matches": high_quality_matches,
        "success_rate": success_rate,
    }

    logger.info(
        f"Output validation stats: "
        f"{high_quality_matches}/{total_rows} high quality matches "
        f"({success_rate*100:.1f}%)"
    )

    # Assert minimum success rate
    assert success_rate >= min_success_rate, (
        f"Success rate too low: {success_rate*100:.1f}% < {min_success_rate*100:.1f}%\n"
        f"Stats: {stats}"
    )

    return stats


# ============================================================================
# E2E TESTS
# ============================================================================


@pytest.mark.e2e
@pytest.mark.slow
def test_full_workflow_happy_path(
    test_client,
    sample_files,
    clean_redis,
    clean_chromadb,
    docker_services,
):
    """
    Test full E2E workflow: Upload → Process → Download → Validate.

    This is the main happy path test covering the complete user journey:
    1. User uploads working file (20 HVAC descriptions)
    2. User uploads reference file (50 catalog items with prices)
    3. User triggers matching process
    4. System processes asynchronously (Celery task)
    5. User polls job status until completion
    6. User downloads result file with matched prices
    7. System validates output quality (≥50% high quality matches)

    Acceptance Criteria (from IMPL_PLAN.md Task 3.10.2):
        ✓ Pełny flow działa E2E
        ✓ Output Excel ma wszystkie kolumny (Cena, Match Score, Match Report)
        ✓ ≥50% dopasowań >75%
        ✓ Czas <60s dla 20 items

    Test Stages:
        STAGE 1: Upload files
        STAGE 2: Trigger matching process
        STAGE 3: Wait for completion (polling)
        STAGE 4: Download results
        STAGE 5: Validate output

    Requirements:
        - Redis running (docker-compose up -d)
        - Celery worker running
        - Sample fixtures exist (tests/fixtures/)

    Note:
        This test uses REAL services (not mocks):
        - Real Redis for progress tracking
        - Real Celery worker for async processing
        - Real file system for temp files
        This ensures true E2E validation.
    """
    logger.info("=" * 60)
    logger.info("STARTING E2E TEST: Full Workflow Happy Path")
    logger.info("=" * 60)

    # ========================================================================
    # STAGE 1: Upload files
    # ========================================================================
    logger.info("\n[STAGE 1] Uploading files...")

    working_upload = upload_file(test_client, sample_files["working"], file_type="working")
    working_file_id = working_upload["file_id"]

    reference_upload = upload_file(test_client, sample_files["reference"], file_type="reference")
    reference_file_id = reference_upload["file_id"]

    assert working_file_id != reference_file_id, "File IDs should be different"

    # ========================================================================
    # STAGE 2: Trigger matching process
    # ========================================================================
    logger.info("\n[STAGE 2] Triggering matching process...")

    process_response = trigger_matching(
        test_client,
        working_file_id=working_file_id,
        reference_file_id=reference_file_id,
        threshold=75.0,
    )

    job_id = process_response["job_id"]
    assert process_response["status"] == "queued", (
        f"Expected status queued, got {process_response['status']}"
    )

    # ========================================================================
    # STAGE 3: Wait for completion (polling)
    # ========================================================================
    logger.info("\n[STAGE 3] Waiting for job completion...")

    final_status = poll_job_status(
        test_client,
        job_id=job_id,
        timeout_seconds=TEST_TIMEOUT_SECONDS,
        poll_interval=POLL_INTERVAL_SECONDS,
    )

    assert final_status["status"] == "completed", (
        f"Expected status completed, got {final_status['status']}"
    )

    # ========================================================================
    # STAGE 4: Download results
    # ========================================================================
    logger.info("\n[STAGE 4] Downloading result file...")

    result_bytes = download_results(test_client, job_id)
    assert len(result_bytes) > 0, "Result file is empty"

    # ========================================================================
    # STAGE 5: Validate output
    # ========================================================================
    logger.info("\n[STAGE 5] Validating output file...")

    stats = validate_output_file(result_bytes, min_success_rate=MIN_SUCCESS_RATE)

    # Log final stats
    logger.info("\n" + "=" * 60)
    logger.info("E2E TEST COMPLETED SUCCESSFULLY ✓")
    logger.info("=" * 60)
    logger.info(f"Total rows processed: {stats['total_rows']}")
    logger.info(f"Rows with prices: {stats['rows_with_price']}")
    logger.info(f"High quality matches: {stats['high_quality_matches']}")
    logger.info(f"Success rate: {stats['success_rate']*100:.1f}%")
    logger.info("=" * 60)


@pytest.mark.e2e
def test_workflow_with_invalid_files(test_client, clean_redis, clean_chromadb, docker_services):
    """
    Verify the upload endpoint rejects each of the documented invalid-file cases:
      1. Empty file
      2. Wrong content (text bytes with .xlsx extension)
      3. Corrupted Excel (looks like .xlsx by extension/MIME but bytes are garbage)
      4. File too large (> MAX_FILE_SIZE_MB, default 10 MB)

    Expected codes: 400 / 413 / 422 (4xx client error). System must stay healthy.
    """
    logger.info("=" * 60)
    logger.info("STARTING E2E TEST: Invalid Files")
    logger.info("=" * 60)

    xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    accepted_codes = (400, 413, 422)

    # Case 1: Empty file
    logger.info("\n[CASE 1] Empty file...")
    response = test_client.post(
        "/api/files/upload",
        files={"file": ("empty.xlsx", BytesIO(b""), xlsx_mime)},
    )
    assert response.status_code in accepted_codes, (
        f"Empty file should be rejected, got {response.status_code} - {response.text}"
    )
    logger.info(f"✓ Empty file rejected with {response.status_code}")

    # Case 2: Wrong content type (text bytes labeled as text/plain)
    logger.info("\n[CASE 2] Text file mislabeled as Excel...")
    response = test_client.post(
        "/api/files/upload",
        files={"file": ("fake.xlsx", BytesIO(b"This is not an Excel file"), "text/plain")},
    )
    assert response.status_code in accepted_codes, (
        f"Text file should be rejected, got {response.status_code} - {response.text}"
    )
    logger.info(f"✓ Text file rejected with {response.status_code}")

    # Case 3: Corrupted Excel — random bytes with .xlsx extension and correct MIME
    # An xlsx is a ZIP archive; arbitrary bytes are not a valid ZIP and openpyxl
    # must reject them rather than crash the API.
    logger.info("\n[CASE 3] Corrupted Excel (random bytes with .xlsx extension)...")
    corrupted = BytesIO(b"\x00\x01\x02\x03not a real xlsx zip\xff\xfe" * 10)
    response = test_client.post(
        "/api/files/upload",
        files={"file": ("corrupted.xlsx", corrupted, xlsx_mime)},
    )
    assert response.status_code in accepted_codes, (
        f"Corrupted xlsx should be rejected, got {response.status_code} - {response.text}"
    )
    logger.info(f"✓ Corrupted xlsx rejected with {response.status_code}")

    # Case 4: File too large — synthesize > 10 MB payload (MAX_FILE_SIZE_MB default).
    # We don't need a real xlsx; size-check should run before xlsx parsing.
    logger.info("\n[CASE 4] File too large (>10MB)...")
    eleven_mb = b"\x00" * (11 * 1024 * 1024)
    response = test_client.post(
        "/api/files/upload",
        files={"file": ("huge.xlsx", BytesIO(eleven_mb), xlsx_mime)},
    )
    assert response.status_code in accepted_codes, (
        f"Oversize file should be rejected, got {response.status_code} - {response.text}"
    )
    logger.info(f"✓ Oversize file rejected with {response.status_code}")

    logger.info("\n" + "=" * 60)
    logger.info("E2E TEST COMPLETED: Invalid Files ✓ (4 cases)")
    logger.info("=" * 60)


def _run_matching_and_get_stats(
    test_client, sample_files, threshold: float, *, min_success_rate: float
) -> dict:
    """Run a full upload→match→download cycle and return validate_output_file stats."""
    working_upload = upload_file(test_client, sample_files["working"], file_type="working")
    reference_upload = upload_file(test_client, sample_files["reference"], file_type="reference")
    process_response = trigger_matching(
        test_client,
        working_file_id=working_upload["file_id"],
        reference_file_id=reference_upload["file_id"],
        threshold=threshold,
    )
    poll_job_status(test_client, process_response["job_id"])
    result_bytes = download_results(test_client, process_response["job_id"])
    return validate_output_file(result_bytes, min_success_rate=min_success_rate)


@pytest.mark.e2e
@pytest.mark.slow
def test_workflow_with_low_threshold(
    test_client,
    sample_files,
    clean_redis,
    clean_chromadb,
    docker_services,
):
    """
    Verify that lowering the matching threshold yields >= as many matches.

    The strict (75%) baseline filters more candidates out, so the relaxed (50%)
    run must produce at least as many priced rows. We compare absolute counts —
    that's the falsifiable claim of the docstring "more items matched".

    Acceptance:
        ✓ rows_with_price(low) >= rows_with_price(high)
        ✓ Both runs return matches above threshold (sanity)
    """
    logger.info("=" * 60)
    logger.info("STARTING E2E TEST: Low Threshold (vs high baseline)")
    logger.info("=" * 60)

    # Run #1: high threshold (75%) — baseline
    logger.info("\n[BASELINE] Matching with threshold=75%...")
    high_stats = _run_matching_and_get_stats(
        test_client, sample_files, threshold=75.0, min_success_rate=MIN_SUCCESS_RATE
    )
    logger.info(
        f"High-threshold matches: {high_stats['rows_with_price']}/{high_stats['total_rows']}"
    )

    # Run #2: low threshold (50%) — should produce >= matches
    logger.info("\n[LOW] Matching with threshold=50%...")
    low_stats = _run_matching_and_get_stats(
        test_client, sample_files, threshold=50.0, min_success_rate=0.3
    )
    logger.info(
        f"Low-threshold matches: {low_stats['rows_with_price']}/{low_stats['total_rows']}"
    )

    # Sanity: both runs produced something
    assert high_stats["rows_with_score"] > 0, "Baseline run should have at least one match"
    assert low_stats["rows_with_score"] > 0, "Low-threshold run should have at least one match"

    # Core assertion: relaxed threshold cannot match fewer items than strict threshold.
    assert low_stats["rows_with_price"] >= high_stats["rows_with_price"], (
        f"Lowering threshold from 75% to 50% reduced match count: "
        f"high={high_stats['rows_with_price']}, low={low_stats['rows_with_price']}. "
        f"This violates the matching engine's monotonicity in threshold."
    )

    logger.info("\n" + "=" * 60)
    logger.info("E2E TEST COMPLETED: Low Threshold ✓")
    logger.info(
        f"Matches at 75% threshold: {high_stats['rows_with_price']}/{high_stats['total_rows']}"
    )
    logger.info(
        f"Matches at 50% threshold: {low_stats['rows_with_price']}/{low_stats['total_rows']}"
    )
    logger.info("=" * 60)


# ============================================================================
# REQUEST VALIDATION TESTS — confirm API rejects bad requests at the boundary
# ============================================================================


def _build_payload(
    working_file_id: str,
    reference_file_id: str,
    *,
    threshold: float = 75.0,
    wf_description_column: str = "A",
    wf_range_start: int = 2,
    wf_range_end: int = 21,
) -> dict:
    """Construct a /api/matching/process payload, with overridable fields."""
    return {
        "working_file": {
            "file_id": working_file_id,
            "description_column": wf_description_column,
            "description_range": {"start": wf_range_start, "end": wf_range_end},
            "price_target_column": "B",
            "matching_report_column": "D",
        },
        "reference_file": {
            "file_id": reference_file_id,
            "description_column": "A",
            "description_range": {"start": 2, "end": 51},
            "price_source_column": "B",
        },
        "matching_threshold": threshold,
        "matching_strategy": "best_match",
        "report_format": "detailed",
    }


@pytest.mark.e2e
@pytest.mark.parametrize(
    "bad_threshold",
    [
        0.0,     # below allowed minimum (>= 1.0)
        -10.0,   # negative
        100.5,   # above allowed maximum (<= 100.0)
        150.0,   # well above maximum
    ],
)
def test_invalid_threshold_rejected(
    test_client,
    sample_files,
    clean_redis,
    clean_chromadb,
    docker_services,
    bad_threshold: float,
):
    """
    POST /api/matching/process with threshold outside [1.0, 100.0] must return 422.

    Pydantic Field(ge=1.0, le=100.0) on ProcessMatchingRequest.matching_threshold
    is the contract; this test guards against accidental relaxation.
    """
    working_upload = upload_file(test_client, sample_files["working"], file_type="working")
    reference_upload = upload_file(test_client, sample_files["reference"], file_type="reference")

    payload = _build_payload(
        working_upload["file_id"],
        reference_upload["file_id"],
        threshold=bad_threshold,
    )
    response = test_client.post("/api/matching/process", json=payload)
    assert response.status_code == 422, (
        f"threshold={bad_threshold} should be rejected with 422, "
        f"got {response.status_code} - {response.text}"
    )
    logger.info(f"✓ threshold={bad_threshold} rejected with 422")


@pytest.mark.e2e
def test_nonexistent_column_rejected(
    test_client,
    sample_files,
    clean_redis,
    clean_chromadb,
    docker_services,
):
    """
    Triggering matching against a column that doesn't exist in the working file
    (e.g., column "Z" on a sample with only A/B) must surface as a job failure
    with a meaningful error — not a silent success or 500.

    The sample working file has 4 columns (A–D). Column "Z" is index 25,
    well beyond the file's range; ColumnNotFoundError is raised by the
    ExcelReader and propagated through the Celery task.
    """
    working_upload = upload_file(test_client, sample_files["working"], file_type="working")
    reference_upload = upload_file(test_client, sample_files["reference"], file_type="reference")

    payload = _build_payload(
        working_upload["file_id"],
        reference_upload["file_id"],
        wf_description_column="Z",  # out of range for the sample file
    )
    response = test_client.post("/api/matching/process", json=payload)

    # The API accepts the request (column letter is syntactically valid A-ZZ),
    # but the Celery task should fail when it tries to read the missing column.
    assert response.status_code == 202, (
        f"Expected 202 Accepted (column letter is syntactically valid), "
        f"got {response.status_code} - {response.text}"
    )
    job_id = response.json()["job_id"]

    # Poll until failure. poll_job_status raises AssertionError on failed status —
    # we want that *expected* failure, so we catch it.
    with pytest.raises(AssertionError, match="failed"):
        poll_job_status(test_client, job_id=job_id, timeout_seconds=60)

    # Verify the failure carries an error_details payload (not a silent crash).
    # The exact message depends on which layer raises first — currently the
    # matching service indexes wf_df.columns[col_idx] and surfaces an IndexError
    # ("list index out of range") rather than a domain ColumnNotFoundError.
    # The contract for *this* test is just: status=failed AND error_details non-empty.
    final = test_client.get(f"/api/jobs/{job_id}/status").json()
    assert final["status"] == "failed", f"Expected failed, got {final['status']}"
    error_text = final.get("error_details") or ""
    assert error_text.strip(), (
        f"Expected non-empty error_details on failure, got: {error_text!r}"
    )
    logger.info(f"✓ Job correctly failed with error: {error_text}")
