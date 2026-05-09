"""
File Storage Service — local-FS implementation of FileStorageServiceProtocol.

Storage layout (TEMP_DIR defaults to /tmp/fastbidder/):
  - uploads/{file_id}/{original_filename}        # uploaded but not yet processed
  - {job_id}/input/{working|reference}_file.xlsx # job inputs (standardized names)
  - {job_id}/output/result.xlsx                  # job result
"""

import logging
import os
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import Optional
from uuid import UUID

from src.domain.shared.exceptions import (
    ExcelParsingError,
    FileSizeExceededError,
)

logger = logging.getLogger(__name__)


class FileStorageService:
    """
    Manages uploads and per-job temp directories on the local FS.

    Business rules:
      - Max 10 MB per file (configurable via MAX_FILE_SIZE_MB env).
      - Allowed extensions: .xlsx, .xls (configurable via ALLOWED_EXTENSIONS env).
      - Cleanup is manual (cleanup_job / cleanup_old_jobs).
    """

    def __init__(
        self,
        base_dir: Optional[str] = None,
        max_size_mb: Optional[int] = None,
        allowed_extensions: Optional[list[str]] = None,
    ) -> None:
        self.base_dir = Path(base_dir or os.getenv("TEMP_DIR", "/tmp/fastbidder"))
        self.max_size_bytes = (
            (max_size_mb or int(os.getenv("MAX_FILE_SIZE_MB", "10"))) * 1024 * 1024
        )
        extensions_str = os.getenv("ALLOWED_EXTENSIONS", ".xlsx,.xls")
        self.allowed_extensions = allowed_extensions or extensions_str.split(",")

        self.base_dir.mkdir(parents=True, exist_ok=True)

    # ---------- Job-storage Protocol methods ----------

    async def file_exists(self, job_id: UUID, file_type: str) -> bool:
        """Check if `{file_type}_file.xlsx` exists in job_id's input/output dir."""
        file_path = self.get_file_path(job_id, file_type)
        exists = file_path.exists()
        logger.debug(f"Checking file existence: {file_path} -> {exists}")
        return exists

    async def get_file_metadata(self, job_id: UUID, file_type: str) -> dict:
        """Return size/format/timestamps for a job-storage file."""
        if not await self.file_exists(job_id, file_type):
            raise FileNotFoundError(
                f"File not found for job_id={job_id}, file_type={file_type}"
            )

        file_path = self.get_file_path(job_id, file_type)
        stat_result = file_path.stat()

        size = stat_result.st_size
        size_mb = round(size / (1024 * 1024), 2)
        format_ext = file_path.suffix.lstrip(".")
        created_at = datetime.fromtimestamp(stat_result.st_ctime).isoformat()
        modified_at = datetime.fromtimestamp(stat_result.st_mtime).isoformat()

        logger.debug(f"Retrieved metadata for {file_type}: {size_mb:.2f}MB")

        return {
            "size": size,
            "size_mb": size_mb,
            "format": format_ext,
            "exists": True,
            "created_at": created_at,
            "modified_at": modified_at,
            "file_type": file_type,
            "file_path": str(file_path),
        }

    def _validate_extension(self, filename: str) -> bool:
        return any(filename.lower().endswith(ext) for ext in self.allowed_extensions)

    def _validate_size(self, file_data: bytes) -> bool:
        return len(file_data) <= self.max_size_bytes

    def _get_job_dir(self, job_id: UUID) -> Path:
        return self.base_dir / str(job_id)

    async def upload_file(
        self,
        job_id: UUID,
        file_data: bytes,
        filename: str,
        file_type: str,
    ) -> Path:
        """Save raw bytes to {job_id}/input/{file_type}_file.xlsx (standardized name)."""
        if not self._validate_extension(filename):
            raise ValueError(
                f"Invalid extension: {filename}. Allowed: {self.allowed_extensions}"
            )
        if not self._validate_size(file_data):
            raise FileSizeExceededError(
                f"File size {len(file_data)} bytes exceeds "
                f"maximum allowed size {self.max_size_bytes} bytes"
            )

        subdir = self._get_subdirectory(file_type)
        target_filename = self._get_filename_for_type(file_type)
        dir_path = self.base_dir / str(job_id) / subdir
        file_path = dir_path / target_filename

        self._ensure_directory_exists(dir_path)
        file_path.write_bytes(file_data)
        self._set_permissions(file_path, mode=0o644)

        logger.info(f"Uploaded {file_type} file: {file_path} ({len(file_data)} bytes)")
        return file_path

    def get_file_path(self, job_id: UUID, file_type: str) -> Path:
        """Return Path to {job_id}/{input|output}/{standardized_filename}. No existence check."""
        subdir = self._get_subdirectory(file_type)
        filename = self._get_filename_for_type(file_type)
        return self.base_dir / str(job_id) / subdir / filename

    async def cleanup_job(self, job_id: UUID) -> None:
        """Hard-delete job_id's directory. Logs warning (not raises) if missing."""
        job_dir = self._get_job_dir(job_id)
        if not job_dir.exists():
            logger.warning(f"Job directory not found for cleanup: {job_id}")
            return
        shutil.rmtree(job_dir)
        logger.info(f"Cleaned up job directory: {job_id}")

    async def cleanup_old_jobs(self, hours: int = 24) -> int:
        """Delete job dirs whose mtime is older than `hours`. Skips `uploads` and non-UUID names."""
        cleaned_count = 0

        if not self.base_dir.exists():
            logger.warning(f"Base directory does not exist: {self.base_dir}")
            return 0

        for job_dir in self.base_dir.iterdir():
            if not job_dir.is_dir():
                continue
            if job_dir.name == "uploads":
                continue
            if not self._is_directory_old(job_dir, hours):
                continue

            try:
                job_id = UUID(job_dir.name)
                await self.cleanup_job(job_id)
                cleaned_count += 1
            except ValueError:
                logger.warning(f"Skipping non-UUID directory: {job_dir.name}")
            except OSError as e:
                logger.error(f"Failed to cleanup {job_dir.name}: {e}")

        logger.info(f"Cleaned up {cleaned_count} jobs older than {hours} hours")
        return cleaned_count

    # ---------- Upload-storage methods ----------

    def get_uploaded_file_path(self, file_id: UUID) -> Path:
        """Return uploads/{file_id}/ directory path. Caller lists/globs for actual file."""
        return self.base_dir / "uploads" / str(file_id)

    async def save_uploaded_file(
        self, file_id: UUID, file_data: bytes, filename: str
    ) -> Path:
        """Save uploaded bytes to uploads/{file_id}/{filename} (preserves original filename)."""
        if not self._validate_extension(filename):
            raise ValueError(
                f"Invalid extension: {filename}. Allowed: {self.allowed_extensions}"
            )
        if not self._validate_size(file_data):
            raise FileSizeExceededError(
                f"File size {len(file_data)} bytes exceeds "
                f"maximum allowed size {self.max_size_bytes} bytes"
            )

        upload_dir = self.base_dir / "uploads" / str(file_id)
        self._ensure_directory_exists(upload_dir)
        file_path = upload_dir / filename
        file_path.write_bytes(file_data)
        self._set_permissions(file_path, mode=0o644)

        logger.info(
            f"Saved uploaded file: {filename} ({len(file_data)} bytes) "
            f"to uploads/{file_id}/"
        )
        return file_path

    async def extract_file_metadata(self, file_path: Path) -> dict:
        """Read structural metadata from an .xlsx (sheets/rows/columns of first sheet)."""
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        stat_result = file_path.stat()
        size = stat_result.st_size
        size_mb = round(size / (1024 * 1024), 2)
        created_at = datetime.fromtimestamp(stat_result.st_ctime).isoformat()
        filename = file_path.name

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            first_sheet = wb.worksheets[0]
            sheets_count = len(wb.sheetnames)
            rows_count = first_sheet.max_row
            columns_count = first_sheet.max_column
            wb.close()
        except Exception as e:
            raise ExcelParsingError(f"Failed to parse Excel file: {e}")

        if sheets_count == 0:
            raise ValueError("Excel file has no sheets")

        logger.debug(
            f"Extracted metadata: {filename} - {sheets_count} sheets, "
            f"{rows_count}x{columns_count}, {size_mb:.2f}MB"
        )

        return {
            "filename": filename,
            "size": size,
            "size_mb": size_mb,
            "sheets_count": sheets_count,
            "rows_count": rows_count,
            "columns_count": columns_count,
            "created_at": created_at,
        }

    async def extract_file_preview(
        self, file_path: Path, rows: int = 5
    ) -> list[dict]:
        """Read first N data rows of the first sheet as list[dict] (header row → keys)."""
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter

            wb = load_workbook(file_path, read_only=True, data_only=True)
            first_sheet = wb.worksheets[0]
            rows_iter = first_sheet.iter_rows(values_only=True)

            header_row = next(rows_iter, None)
            if header_row is None:
                wb.close()
                logger.debug(f"Extracted preview: 0 rows from {file_path.name} (empty file)")
                return []

            # None-valued header cells → fallback to Excel column letter (A, B, …)
            headers = []
            for idx, cell_value in enumerate(header_row):
                if cell_value is not None:
                    headers.append(str(cell_value))
                else:
                    headers.append(get_column_letter(idx + 1))

            preview_rows = []
            for row_idx, row_values in enumerate(rows_iter):
                if row_idx >= rows:
                    break
                row_dict = {}
                for col_idx, cell_value in enumerate(row_values):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = cell_value
                preview_rows.append(row_dict)

            wb.close()
        except Exception as e:
            raise ExcelParsingError(f"Failed to read Excel file: {e}")

        if len(preview_rows) == 0:
            logger.debug(f"Extracted preview: 0 rows from {file_path.name} (no data rows)")
            return []

        logger.debug(f"Extracted preview: {len(preview_rows)} rows from {file_path.name}")
        return preview_rows

    # ---------- Result-file methods ----------

    def get_result_file_path(self, job_id: UUID) -> Path:
        """Return path to {job_id}/output/result.xlsx (no existence check)."""
        result_path = self.get_file_path(job_id, "result")
        logger.debug(f"Result file path for job {job_id}: {result_path}")
        return result_path

    def result_file_exists(self, job_id: UUID) -> bool:
        """Quick existence check for the result file."""
        result_path = self.get_result_file_path(job_id)
        exists = result_path.exists()
        logger.debug(f"Result file exists check for job {job_id}: {exists}")
        return exists

    # ---------- Internal helpers ----------

    def _get_subdirectory(self, file_type: str) -> str:
        if file_type in ("working", "reference"):
            return "input"
        elif file_type == "result":
            return "output"
        else:
            raise ValueError(
                f"Unknown file_type: '{file_type}'. "
                f"Expected 'working', 'reference', or 'result'."
            )

    def _get_filename_for_type(self, file_type: str) -> str:
        mapping = {
            "working": "working_file.xlsx",
            "reference": "reference_file.xlsx",
            "result": "result.xlsx",
        }
        if file_type not in mapping:
            raise ValueError(
                f"Unknown file_type: '{file_type}'. "
                f"Expected 'working', 'reference', or 'result'."
            )
        return mapping[file_type]

    def _ensure_directory_exists(self, dir_path: Path) -> None:
        if dir_path.exists():
            return
        dir_path.mkdir(parents=True, exist_ok=True)
        self._set_permissions(dir_path, mode=0o755)
        logger.debug(f"Created directory: {dir_path}")

    def _set_permissions(self, path: Path, mode: int) -> None:
        # Windows NTFS / FAT32 may not support chmod — log and continue rather than fail.
        try:
            os.chmod(path, mode)
            logger.debug(f"Set permissions {oct(mode)} on {path}")
        except (OSError, NotImplementedError):
            logger.debug(f"Could not set permissions on {path} (Windows or unsupported filesystem)")

    def _atomic_write_file(self, file_path: Path, data: bytes) -> None:
        """Write to {file_path}.tmp then rename — guarantees no partially-written file on crash."""
        tmp_path = file_path.with_suffix(file_path.suffix + ".tmp")
        tmp_path.write_bytes(data)
        self._set_permissions(tmp_path, mode=0o644)
        # replace() (not rename()) to overwrite atomically on Windows.
        tmp_path.replace(file_path)
        logger.debug(f"Atomic write: {len(data)} bytes to {file_path}")

    def _is_directory_old(self, dir_path: Path, hours: int) -> bool:
        """Check if directory's mtime is older than `hours` ago. mtime > ctime since it tracks content changes."""
        if not dir_path.exists():
            return False
        mtime = dir_path.stat().st_mtime
        cutoff_timestamp = time.time() - (hours * 3600)
        return mtime < cutoff_timestamp
