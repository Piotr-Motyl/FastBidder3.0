"""
Excel writer — writes matching results (prices, reports, coloring) to Excel files using openpyxl.
"""

import shutil
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.application.models import ReportFormat
from src.domain.hvac.entities.hvac_description import HVACDescription


class ExcelWriterService:
    """
    Writes matched prices and reports into the original working Excel (preserves formatting).

    New columns added: "Cena" (price), "Match Score" (0-100), "Confidence" (0-1), "Match Report".
    Conditional coloring: Green (>90%), Yellow (75-90%), Red (<75%).
    """

    # Color definitions for conditional formatting (RGB hex)
    COLOR_GREEN = "00FF00"  # Score > 90%
    COLOR_YELLOW = "FFFF00"  # Score 75-90%
    COLOR_RED = "FF0000"  # Score < 75%

    def __init__(self) -> None:
        pass

    @staticmethod
    def _column_letter_to_index(column: str) -> int:
        """
        Convert Excel column letter to 1-based index (A=1, B=2, Z=26, AA=27, base-26).

        Raises ValueError for non-alphabetic input.
        """
        # Validate input
        if not column or not column.isalpha():
            raise ValueError(f"Column must contain only letters, got: '{column}'")

        # Convert Excel column letter to 1-based index (A=1, B=2, AA=27, etc.)
        index = 0
        for char in column.upper():
            index = index * 26 + (ord(char) - ord('A') + 1)

        return index

    def write_results_to_file(
        self,
        original_file_path: Path,
        descriptions: list[HVACDescription],
        price_column: str,
        report_column: Optional[str] = None,
        output_path: Optional[Path] = None,
        report_format: ReportFormat = ReportFormat.SIMPLE,
        sheet_name: Optional[str] = None,
    ) -> Path:
        """
        Write prices and reports into original Excel file using openpyxl (preserves formatting).

        Backup → load workbook → write prices → apply coloring → write reports (if report_column)
            → auto-size columns → save to output_path.

        Coloring (price_column): Green >90%, Yellow 75-90%, Red <75%, no color if no match.
        output_path defaults to original_file_path.parent / "result.xlsx".
        Raises FileNotFoundError, ValueError, OSError.
        """
        # Step 1: Create backup of original file
        self._create_backup(original_file_path)

        # Step 2: Load workbook with openpyxl (preserves formatting)
        workbook = self._load_workbook(original_file_path)

        # Step 3: Get worksheet (first sheet or specified sheet)
        worksheet = self._get_worksheet(workbook, sheet_name)

        # Step 4: Write prices to price_column
        self._write_prices_to_column(worksheet, descriptions, price_column)

        # Step 5: Apply conditional coloring to price cells
        self._apply_cell_coloring(worksheet, descriptions, price_column)

        # Step 6: Write match reports to report_column (if specified)
        if report_column:
            self._write_reports_to_column(
                worksheet, descriptions, report_column, report_format
            )

        # Step 7: Auto-size columns for readability
        columns_to_resize = [price_column]
        if report_column:
            columns_to_resize.append(report_column)
        self._autosize_columns(worksheet, columns_to_resize)

        # Step 8: Determine output path (default: parent / "result.xlsx")
        if output_path is None:
            output_path = original_file_path.parent / "result.xlsx"

        # Step 9: Save workbook to output_path
        result_path = self._save_workbook(workbook, output_path)

        return result_path

    async def write_results(
        self,
        descriptions: list[HVACDescription],
        output_path: Path,
        original_path: Optional[Path] = None,
    ) -> Path:
        """Not implemented. Use write_results_to_file() instead."""
        raise NotImplementedError(
            "write_results() to be implemented in Phase 3. "
            "Will create Polars DataFrame from descriptions and write to Excel."
        )

    def _create_backup(self, original_path: Path) -> Path:
        """Copy file to {stem}_backup{suffix} in same directory. Raises FileNotFoundError if missing."""
        # Check if original file exists
        if not original_path.exists():
            raise FileNotFoundError(f"File not found: {original_path}")

        # Create backup path: parent directory / filename_backup.extension
        backup_path = original_path.parent / f"{original_path.stem}_backup{original_path.suffix}"

        # Copy file with metadata preservation (shutil.copy2 preserves timestamps)
        shutil.copy2(original_path, backup_path)

        return backup_path

    def _load_workbook(self, file_path: Path) -> Workbook:
        """Load workbook with openpyxl (data_only=False preserves formulas). Raises FileNotFoundError."""
        # Check if file exists
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Load workbook with openpyxl
        # data_only=False preserves formulas (not just calculated values)
        workbook = load_workbook(filename=file_path, data_only=False)

        return workbook

    def _get_worksheet(
        self, workbook: Workbook, sheet_name: Optional[str] = None
    ) -> Worksheet:
        """Return active sheet if sheet_name is None, else named sheet. Raises ValueError if not found."""
        # If sheet_name not specified, use active sheet (first sheet)
        if sheet_name is None:
            worksheet = workbook.active
            if worksheet is None:
                raise ValueError("Workbook has no active sheet")
            return worksheet

        # Get worksheet by name
        try:
            return workbook[sheet_name]
        except KeyError:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available sheets: {available_sheets}"
            )

    def _write_prices_to_column(
        self,
        worksheet: Worksheet,
        descriptions: list[HVACDescription],
        column: str,
    ) -> None:
        """Write desc.matched_price to column at desc.source_row_number. Skips None prices."""
        # Convert column letter to 1-based index for openpyxl
        column_index = self._column_letter_to_index(column)

        # Write prices to column for each description
        for desc in descriptions:
            # Skip rows without a matched price or without a known target row
            if desc.matched_price is None or desc.source_row_number is None:
                continue

            cell = worksheet.cell(row=desc.source_row_number, column=column_index)
            cell.value = float(desc.matched_price)

    def _write_reports_to_column(
        self,
        worksheet: Worksheet,
        descriptions: list[HVACDescription],
        column: str,
        report_format: ReportFormat,
    ) -> None:
        """Write desc.get_match_report() to column at desc.source_row_number. Skips None scores."""
        # Convert column letter to 1-based index for openpyxl
        column_index = self._column_letter_to_index(column)

        # Write reports to column for each description
        for desc in descriptions:
            if desc.match_score is None or desc.source_row_number is None:
                continue

            report = desc.get_match_report()
            if not report:
                continue

            cell = worksheet.cell(row=desc.source_row_number, column=column_index)
            cell.value = report

    def _apply_cell_coloring(
        self,
        worksheet: Worksheet,
        descriptions: list[HVACDescription],
        column: str,
    ) -> None:
        """Apply PatternFill to price_column cells based on match score thresholds."""
        # Convert column letter to 1-based index for openpyxl
        column_index = self._column_letter_to_index(column)

        # Apply coloring to cells for each description
        for desc in descriptions:
            if (
                desc.matched_price is None
                or desc.match_score is None
                or desc.source_row_number is None
            ):
                continue

            color = self._get_color_for_score(desc.match_score.final_score)
            if not color:
                continue

            cell = worksheet.cell(row=desc.source_row_number, column=column_index)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _get_color_for_score(self, score: float) -> Optional[str]:
        """Return RGB hex: Green (>90%), Yellow (75-90%), Red (<75%)."""
        if score > 90:
            return self.COLOR_GREEN
        elif score >= 75:
            return self.COLOR_YELLOW
        else:
            return self.COLOR_RED

    def _autosize_columns(
        self, worksheet: Worksheet, columns: list[str]
    ) -> None:
        """Set each column's width to max cell content length + 2 (minimum 8)."""
        # Auto-size each column in the list
        for column_letter in columns:
            # Calculate maximum content width for this column
            max_width = 0

            # Iterate through all cells in this column
            for cell in worksheet[column_letter]:
                # Skip empty cells
                if cell.value is None:
                    continue

                # Calculate length of cell value as string
                cell_length = len(str(cell.value))

                # Update max width
                max_width = max(max_width, cell_length)

            # Set column width (add padding of 2 for readability)
            # Minimum width of 8 to ensure columns are visible
            adjusted_width = max(max_width + 2, 8)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_dataframe_to_excel(self, dataframe, output_path: Path) -> Path:
        """Save Polars or Pandas DataFrame to Excel via openpyxl. Creates parent dirs if missing."""
        # Create parent directory if doesn't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Convert to Pandas if needed (Polars has to_pandas() method)
        if hasattr(dataframe, 'to_pandas'):
            # Polars DataFrame
            pandas_df = dataframe.to_pandas()
        else:
            # Already Pandas DataFrame
            pandas_df = dataframe

        # Save to Excel using openpyxl
        pandas_df.to_excel(output_path, index=False, engine="openpyxl")

        return output_path

    def _save_workbook(self, workbook: Workbook, output_path: Path) -> Path:
        """Save workbook to output_path. Creates parent dirs. Returns output_path."""
        # Create parent directory if it doesn't exist
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook to file
        workbook.save(output_path)

        return output_path

    async def write_unmatched_report(
        self, descriptions: list[HVACDescription], output_path: Path
    ) -> Path:
        """Not implemented."""
        raise NotImplementedError(
            "write_unmatched_report() to be implemented in Phase 3. "
            "Will create Excel report of items requiring manual review."
        )
